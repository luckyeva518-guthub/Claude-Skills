"""
CDQ v3 — JSON→Excel 字段解析模板（通用版）

使用方式：
  1. 准备 backend_data（从 Catalog API JSON 解析后的 dict）
  2. 准备 frontend_rows（从前台 web_reader 提取的 [(field, value)] 列表）
  3. 调用 generate_excel(asin, backend_data, frontend_rows, missing_rows, output_path)

依赖：pip install openpyxl
"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def load_field_mapping(mapping_path=None):
    """加载 field-mapping.json 映射表"""
    if mapping_path is None:
        import os
        mapping_path = os.path.join(os.path.dirname(__file__), "field-mapping.json")
    with open(mapping_path, "r", encoding="utf-8") as f:
        mapping_data = json.load(f)
    return mapping_data.get("fields", {})


def get_field_info(field_map, dot_key):
    """获取字段的中文名和枚举值映射"""
    info = field_map.get(dot_key, None)
    if info:
        cn_name = info.get("cn", dot_key)
        enum_vals = info.get("values", None)
        return cn_name, enum_vals
    return dot_key, None


def translate_enum_value(enum_vals, val):
    """翻译枚举值为 'Original（中文）' 格式"""
    if enum_vals and val in enum_vals:
        return f"{val}（{enum_vals[val]}）"
    return val


def format_value(value, enum_vals=None):
    """格式化字段值用于 Excel 显示"""
    if value is None:
        return ""
    elif isinstance(value, bool):
        return str(value)
    elif isinstance(value, (int, float)):
        return str(value)
    elif isinstance(value, str):
        return translate_enum_value(enum_vals, value) if enum_vals else value
    elif isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                parts.append(json.dumps(item, ensure_ascii=False))
            else:
                parts.append(str(item))
        return "\n".join(parts)
    elif isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    else:
        return str(value)


def add_row(rows, field_map, key, value, source="后台API"):
    """添加一行数据到 rows 列表"""
    cn_name, enum_vals = get_field_info(field_map, key)
    field_display = f"{key}（{cn_name}）" if cn_name != key else key
    val_display = format_value(value, enum_vals)
    rows.append((field_display, val_display, source))


def extract_backend_data(data, field_map):
    """从 Catalog API JSON 提取后台数据，返回 [(field, value, source)] 列表"""
    rows = []

    # 基础字段
    add_row(rows, field_map, "asin", data.get("asin"), "后台API")
    rows.append(("imageUrl（缩略图 URL）", data.get("imageUrl", ""), "后台API"))
    add_row(rows, field_map, "product_type", data.get("productType"), "后台API")

    # 标题
    item_name = None
    if data.get("itemName") and data["itemName"].get("value"):
        item_name = data["itemName"]["value"]
    add_row(rows, field_map, "title", item_name, "后台API")

    # 从 reconciledListing.detailPageListingResponse 提取
    dplr = None
    rlisting = data.get("reconciledListing")
    if rlisting:
        dplr = rlisting.get("detailPageListingResponse")
    if dplr:
        field_extractors = [
            ("brand#1.value", "attributes.brand"),
            ("manufacturer#1.value", "attributes.manufacturer"),
            ("product_type#1.value", "product_type"),
            ("variation_theme#1.name", "attributes.variation_theme"),
            ("product_site_launch_date#1.value", "attributes.launch_date"),
        ]
        for api_key, field_key in field_extractors:
            obj = dplr.get(api_key, {})
            if isinstance(obj, dict) and "value" in obj:
                add_row(rows, field_map, field_key, obj["value"], "后台API")

        # item_type_keyword（特殊处理，保留自定义显示名）
        itk = dplr.get("item_type_keyword#1.value", {})
        if isinstance(itk, dict) and "value" in itk:
            rows.append(("item_type_keyword（物品类型关键词）", itk["value"], "后台API"))

        # sales_rank
        sr = dplr.get("sales_rank", {})
        if isinstance(sr, dict) and "value" in sr:
            add_row(rows, field_map, "sales_rank.main", sr["value"], "后台API")

        # 图片
        primary_img = dplr.get("detail_page_primary_image_url", {})
        if isinstance(primary_img, dict) and "value" in primary_img:
            add_row(rows, field_map, "images.primary", primary_img["value"], "后台API")

        img_count = 1
        for key in [f"pt{i:02d}_image_url" for i in range(1, 9)]:
            img_obj = dplr.get(key, {})
            if isinstance(img_obj, dict) and "value" in img_obj:
                img_count += 1
                rows.append((f"images.{key}（图片 {key}）", img_obj["value"], "后台API"))
        add_row(rows, field_map, "images.count", img_count, "后台API")

    # 从 detailPageSummary 提取
    dps = data.get("detailPageSummary")
    if dps:
        sales_ranks = dps.get("salesRanks", [])
        for sr in sales_ranks:
            ctx = sr.get("storeContextName", "")
            rank = sr.get("rank", "")
            # 判断子类目 vs 主类目
            display_name = "sales_rank.sub（子类目 BSR 排名）" if rank != dps.get("salesRanks", [{}])[0].get("rank") else "sales_rank.main（主类目 BSR 排名）"
            rows.append((display_name, f"#{rank} ({ctx})", "后台API"))

        crs = dps.get("customerReviewSummary", {})
        if crs:
            add_row(rows, field_map, "reviews.count", crs.get("reviewCount"), "后台API")
            add_row(rows, field_map, "reviews.rating", crs.get("reviewStars"), "后台API")

        offer_summaries = dps.get("offerSummaries", [])
        for os_item in offer_summaries:
            condition = os_item.get("offerType", "")
            price = os_item.get("lowestOfferPrice", "")
            num = os_item.get("numOffers", "")
            rows.append((f"price（价格 - {condition}）", f"{price} ({num} offers)", "后台API"))

        dpl = dps.get("detailPageLink", "")
        if dpl:
            rows.append(("detailPageLink（详情页链接）", dpl, "后台API"))

    return rows


def generate_excel(asin, backend_data, frontend_rows=None, missing_rows=None, output_path=None):
    """
    生成 CDQ 字段解析 Excel 文件

    Args:
        asin: ASIN 字符串（如 "B0GSFCFRPM"）
        backend_data: Catalog API 返回的 JSON dict
        frontend_rows: 前台验证数据 [(field_display, value, "前台验证"), ...]
        missing_rows: 缺失/未确认字段 [(field_display, warning, "—"), ...]
        output_path: 输出文件路径，默认为桌面上的 {ASIN}_字段解析.xlsx

    Returns:
        output_path: 生成的 Excel 文件路径
    """
    if output_path is None:
        import os
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        output_path = os.path.join(desktop, f"{asin}_字段解析.xlsx")

    field_map = load_field_mapping()

    # 提取后台数据
    all_rows = extract_backend_data(backend_data, field_map)

    # 添加前台验证数据
    if frontend_rows:
        all_rows.extend(frontend_rows)

    # 添加缺失/未确认字段
    if missing_rows:
        all_rows.extend(missing_rows)

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{asin} 字段解析"

    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )
    wrap_align = Alignment(vertical="top", wrap_text=True)
    frontend_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    section_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    section_font = Font(bold=True, size=11, color="1E3A5F")

    # 标题行
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = f"{asin} — 后台数据 + 前台验证 字段解析"
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头行
    for col, header in enumerate(["字段", "值", "数据来源"], 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row = 3

    def write_section_header(title):
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        ws.merge_cells(f"A{current_row}:C{current_row}")
        for c in range(1, 4):
            ws.cell(row=current_row, column=c).border = thin_border
            ws.cell(row=current_row, column=c).fill = section_fill
        current_row += 1

    def write_data_row(field, val, source):
        nonlocal current_row
        cell_a = ws.cell(row=current_row, column=1, value=field)
        cell_b = ws.cell(row=current_row, column=2, value=val)
        cell_c = ws.cell(row=current_row, column=3, value=source)
        for cell in [cell_a, cell_b, cell_c]:
            cell.border = thin_border
        cell_a.alignment = wrap_align
        cell_b.alignment = wrap_align
        cell_c.alignment = Alignment(horizontal="center", vertical="top")

        if isinstance(val, str) and val.startswith("http"):
            cell_b.font = Font(color="2563EB", underline="single")
        if isinstance(val, str) and "⚠️" in val:
            cell_b.fill = warn_fill
            cell_b.font = Font(color="D97706")
        elif source == "前台验证":
            cell_a.fill = frontend_fill
            cell_b.fill = frontend_fill
            cell_c.fill = frontend_fill
        current_row += 1

    # 第一部分：后台 API 数据
    backend_rows = [r for r in all_rows if r[2] == "后台API"]
    if backend_rows:
        write_section_header("【后台 Catalog API 数据】")
        for field, val, source in backend_rows:
            write_data_row(field, val, source)

    # 第二部分：前台验证数据
    frontend_data = [r for r in all_rows if r[2] == "前台验证"]
    if frontend_data:
        write_section_header("【前台 Product Information 验证数据】")
        for field, val, source in frontend_data:
            write_data_row(field, val, source)

    # 第三部分：缺失/未确认字段
    missing_data = [r for r in all_rows if r[2] not in ("后台API", "前台验证")]
    if missing_data:
        write_section_header("【缺失/未确认字段】")
        for field, val, source in missing_data:
            write_data_row(field, val, source)

    # 列宽
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 85
    ws.column_dimensions["C"].width = 14

    ws.freeze_panes = "A3"

    wb.save(output_path)
    print(f"Excel saved to: {output_path}")
    print(f"Total data rows: {len(all_rows)}")
    return output_path


# ============================================================
# 直接运行时的示例用法
# ============================================================
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel-template.py <backend_json_path> [output_path]")
        print("  backend_json_path: Catalog API 返回的 JSON/TXT 文件路径")
        print("  output_path: 输出 Excel 路径（默认：桌面/{ASIN}_字段解析.xlsx）")
        sys.exit(1)

    json_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else None

    # 读取并解析 JSON
    with open(json_path, "r", encoding="utf-8") as f:
        raw = f.read()

    line = raw.strip()
    if "\t" in line:
        json_str = line.split("\t", 1)[1]
    else:
        json_str = line
    data = json.loads(json_str)

    asin = data.get("asin", "UNKNOWN")

    # 仅后台数据，前台数据需要在 CDQ 诊断流程中由 AI 动态填充
    generate_excel(asin, data, frontend_rows=None, missing_rows=None, output_path=out_path)
